from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import openpyxl, os

# EXCEL SETUP
excel_file = "personal_details_clean.xlsx"

# Create new Excel file if it doesn't exist
if not os.path.exists(excel_file):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Personal Info"
    headers = [
        "Aadhar Number", "Name", "Father's Name", "Mother's Name", 
        "Date Of Birth", "Age", "Gender", "Caste", "Religion",
        "Qualification", "Email ID", "Mobile No", "Address"
    ]
    ws.append(headers)
    wb.save(excel_file)

# TKINTER UI
window = Tk()
window.title("Personal Details Form")
window.geometry("600x800")
window.configure(bg="#dde6ed")
window.resizable(False, False)

header_font = ("Segoe UI", 18, "bold")
label_font = ("Segoe UI", 12)
entry_font = ("Segoe UI", 12)

ttk.Style().theme_use('clam')
ttk.Style().configure("TLabel", background="#dde6ed", font=label_font)
ttk.Style().configure("TButton", font=label_font)

ttk.Label(window, text="📝 Personal Details Form", font=header_font, background="#dde6ed", foreground="#2c3e50").pack(pady=20)

form_frame = Frame(window, bg="#dde6ed")
form_frame.pack(pady=10)

fields = [
    "Aadhar Number", "Name", "Father's Name", "Mother's Name", 
    "Date Of Birth (DD-MM-YYYY)", "Age", "Gender",
    "Caste", "Religion", "Qualification", "Email ID", "Mobile No"
]

entries = {}

for idx, field in enumerate(fields):
    Label(form_frame, text=field, bg="#dde6ed", font=label_font).grid(row=idx, column=0, sticky=W, padx=20, pady=6)
    if field == "Gender":
        var = StringVar()
        cb = ttk.Combobox(form_frame, textvariable=var, values=["Male", "Female", "Other"], state="readonly", width=28, font=entry_font)
        cb.grid(row=idx, column=1, padx=10, pady=6)
        entries[field] = cb
    else:
        ent = Entry(form_frame, width=30, font=entry_font)
        ent.grid(row=idx, column=1, padx=10, pady=6)
        entries[field] = ent

Label(form_frame, text="Address", bg="#dde6ed", font=label_font).grid(row=len(fields), column=0, sticky=W, padx=20, pady=6)
address = Text(form_frame, width=30, height=4, font=entry_font)
address.grid(row=len(fields), column=1, padx=10, pady=6)

# FUNCTIONS
def clear_form():
    for w in entries.values():
        if isinstance(w, ttk.Combobox):
            w.set("")
        else:
            w.delete(0, END)
    address.delete("1.0", END)

def submit_form():
    data = [w.get().strip() for w in entries.values()]
    if "" in data or not address.get("1.0", END).strip():
        messagebox.showwarning("Error", "All fields are required.")
        return
    data.append(address.get("1.0", END).strip())
    wb = openpyxl.load_workbook(excel_file)
    wb.active.append(data)
    wb.save(excel_file)
    messagebox.showinfo("Success", "✅ Data saved!")
    clear_form()

def search_data():
    aadhar = entries["Aadhar Number"].get().strip()
    if not aadhar:
        messagebox.showwarning("Warning", "Enter Aadhar Number to search.")
        return
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == aadhar:
            result = "\n".join(f"{fields[i]}: {row[i]}" for i in range(len(fields))) + f"\nAddress: {row[-1]}"
            messagebox.showinfo("Found", result)
            return
    messagebox.showinfo("Not Found", "❌ Record not found.")

def delete_data():
    aadhar = entries["Aadhar Number"].get().strip()
    if not aadhar:
        messagebox.showwarning("Warning", "Enter Aadhar Number to delete.")
        return
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == aadhar:
            ws.delete_rows(row)
            wb.save(excel_file)
            messagebox.showinfo("Deleted", "✅ Record deleted.")
            clear_form()
            return
    messagebox.showinfo("Not Found", "❌ No matching record found.")

# BUTTONS
btn_frame = Frame(window, bg="#dde6ed")
btn_frame.pack(pady=20)

def custom_btn(master, text, cmd, color):
    btn = Button(master, text=text, command=cmd, width=12, font=("Segoe UI", 11, "bold"),
                 fg="white", bg=color, activebackground=color, bd=0, cursor="hand2")
    btn.pack(side=LEFT, padx=12)
    return btn

custom_btn(btn_frame, "Submit", submit_form, "#27ae60")
custom_btn(btn_frame, "Search", search_data, "#2980b9")
custom_btn(btn_frame, "Delete", delete_data, "#c0392b")
custom_btn(btn_frame, "Clear", clear_form, "#f39c12")

window.mainloop()
